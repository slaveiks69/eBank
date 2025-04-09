from flask import Blueprint, json, request
import yaml
#from index import cfg

export = Blueprint('export', __name__)

import openpyxl, shutil, os, datetime

def sex():
    with open("config.yaml", encoding="UTF-8") as f:
        cfg = yaml.load(f, Loader=yaml.FullLoader)
        return cfg
    
cfg = sex()

@export.post('/export')
def export_post(gexport = False, gdict = None):
    if gexport is False:
        arg = json.loads(request.data)
        persons_to_export = arg['data']
    else:
        persons_to_export = gdict

    shutil.copy(cfg["excel"]["formPath"], cfg["excel"]["copyFormPath"])
    
    workbook = openpyxl.load_workbook(cfg["excel"]["copyFormPath"])
    sheet = workbook['Лист1']

    row_start = 4
    index = 0

    print(persons_to_export)

    for person in persons_to_export:
        if gexport is False:
            birthDate = datetime.datetime.strptime(person['birthDate'],"%a, %d %b %Y %H:%M:%S GMT")
            passportIssueDate = datetime.datetime.strptime(person['passportIssueDate'],"%a, %d %b %Y %H:%M:%S GMT")
        else:
            birthDate = datetime.datetime.strptime(person['birthDate'],"%d.%m.%Y")
            passportIssueDate = person['passportIssueDate']
        

        sheet.cell(row=row_start, column=1).value = index # индексация с 0
        sheet.cell(row=row_start, column=2).value = person['lastName'] # фамилия
        sheet.cell(row=row_start, column=3).value = person['firstName'] # имя
        sheet.cell(row=row_start, column=4).value = person['patronymic'] # отчество
        sheet.cell(row=row_start, column=5).value = birthDate.strftime("%d.%m.%Y") # дата рождения
        sheet.cell(row=row_start, column=6).value = person['birthPlace'] # место рождения
        sheet.cell(row=row_start, column=7).value = person['passportSerial'] # серия и номер паспорта
        sheet.cell(row=row_start, column=8).value = person['passportIssue'] # где и когда выдан
        sheet.cell(row=row_start, column=9).value = passportIssueDate.strftime("%d.%m.%Y") # дата выдачи
        sheet.cell(row=row_start, column=10).value = person['passportDivisionCode'] # код подразделения
        sheet.cell(row=row_start, column=11).value = person['address'] # адрес проживания
        sheet.cell(row=row_start, column=12).value = 'Москва' # город
        sheet.cell(row=row_start, column=13).value = person['phoneMobile'] # домашний телефон
        sheet.cell(row=row_start, column=14).value = person['phoneMobile'] # моб телефон
        sheet.cell(row=row_start, column=15).value = person['codeword'] # код слово
        sheet.cell(row=row_start, column=16).value = 'Тимирязевский' # тимирязевский

        row_start = row_start + 1
        index = index + 1

    if True is True:
        name = ""
        k = 1
        now = datetime.datetime.now().strftime("%Y.%m.%d")

        ExcelServerPath = cfg["default"]["serverPath"]+"\\"+cfg["default"]["folder"]

        while (True):
            name = f"ВТБ_{now}_{k}.xlsx"
            if os.path.exists(ExcelServerPath+cfg["excel"]["savePath"]+name):
                k = k + 1
            else:
                break
        
        workbook.save(ExcelServerPath+cfg["excel"]["savePath"]+name)

        save_folder_path = ExcelServerPath+cfg["excel"]["exportsPath"]+"\\"+now+"\\"

        if not os.path.exists(save_folder_path):
            os.makedirs(save_folder_path)

        name = ""
        k = 1

        while (True):
            name = f"ВТБ_{now}_{k}.xlsx"
            if os.path.exists(save_folder_path+name):
                k = k + 1
            else:
                break 
        
        workbook.save(save_folder_path+name)
    else:
        workbook.save(cfg["excel"]["copyFormPath"])

    return { 'export': 'okay' }