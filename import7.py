from flask import Blueprint, json, render_template, request, redirect, url_for
from openpyxl import load_workbook
from utils import *

#from index import socketio

import7 = Blueprint('import', __name__)

class importPerson:
    def getPassportSerialNumber(self):
        if self.passportSerial != None:
            if self.passportNumber != None:
                return self.passportSerial + " " + self.passportNumber

    def passport_serial(self, ps: str):
        return ps[:2]+" "+ps[2:]
    
    def get_birth_date(self):
        if self.birthDate != None:
            return self.birthDate.strftime("%d.%m.%Y")

    def get_issue_date(self):
        if self.passportIssueDate != None:
            return self.passportIssueDate.strftime("%Y-%m-%d")

    def __init__(self, k, lN, fN, p, bD, tD, pS, pN, pI, pID, pDC, tA, A, card, bP):
        self.kod = k
        self.lastName = lN
        self.firstName = fN
        self.patronymic = p
        self.birthDate = bD
        self.typeDocument = tD
        self.passportSerial = self.passport_serial(str(pS))
        self.passportNumber = pN
        self.passportIssue = pI
        self.passportIssueDate = pID
        self.passportDivisionCode = pDC
        self.trueAdress = tA
        self.adress = A
        self.cardNumber = card
        self.birthPlace = bP

def get_persons_from_sheet(sheet):
    persons = []

    for personExcel in sheet.iter_rows():

        person = importPerson(
            personExcel[0].value,
            personExcel[1].value,
            personExcel[2].value,
            personExcel[3].value,
            personExcel[4].value,
            personExcel[5].value,
            personExcel[6].value,
            personExcel[7].value,
            personExcel[8].value,
            personExcel[9].value,
            personExcel[10].value,
            personExcel[11].value,
            personExcel[12].value,
            personExcel[18].value,
            personExcel[21].value
        )

        persons.append(person)

    return persons

def import_get():
    workbook = load_workbook(static.cfg["default"]["serverPath"]+"\\"+static.cfg["default"]["folder"]+static.cfg["excel"]["semerkiPath"]+static.cfg["excel"]["semerkiFile"])
    sheet = workbook['Лист1']

    sheet.delete_rows(0,1)
    
    #print(get_persons_from_sheet(sheet))

    excelPersons = get_persons_from_sheet(sheet)

    for excelPerson in excelPersons:
        bankPerson = find_in_bank(excelPerson.getPassportSerialNumber())
        if bankPerson != None:
            if excelPerson.lastName != bankPerson['lastName']:
                ...
            if excelPerson.firstName != bankPerson['firstName']:
                ...
            if excelPerson.patronymic != bankPerson['patronymic']:
                ...
            if excelPerson.get_birth_date() != bankPerson['birthDateFormated']:
                ...
            if excelPerson.get_issue_date() != bankPerson['passportIssueDate'].strftime("%Y-%m-%d"):
                ...
            if excelPerson.passportDivisionCode != bankPerson['passportDivisionCode']:
                ...
            if excelPerson.cardNumber != bankPerson['card']:
                sql = f"UPDATE person_card SET account_number = '{excelPerson.cardNumber}' WHERE passport_serial = '{excelPerson.getPassportSerialNumber()}' IF @@ROWCOUNT = 0 INSERT INTO person_card (passport_serial, account_number) VALUES ('{excelPerson.getPassportSerialNumber()}','{excelPerson.cardNumber}');"

            else:
                ...
        else:
            ...
    
    return f"{len(excelPersons)}"

@import7.get('/import')
def imports():
    
    return render_template('import.html', uri=static.uri)
