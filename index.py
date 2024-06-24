import datetime
from flask import Flask, json, render_template, request, redirect, url_for
from utils import *



app = Flask(__name__)

app.config['SERVER_NAME'] = 'localhost:1111'

from monitoring import monitoring

app.register_blueprint(monitoring, url_prefix='/')

print(static.pc_ip)

@app.get('/')
def home():

    #check_local_database()
    #print(get_persons())

    #print(get_recruits_statistic())  

    return render_template('home.html', ver=static.cfg["default"]["version"], bgimg=static.cfg["default"]["background_image"])

@app.get('/test')
def hello_world1():
    return get_persons() 

@app.post('/find')
def find_people():
    arg = json.loads(request.data)
    kod = str(arg['kod'])[:2]+str(arg['kod'])[3:][:2]+str(arg['kod'])[6:]
    found = { 'found': 'none' }

    # bank baze
    if found['found'] == 'none':
        person = find_in_bank(arg['kod'])
        if person != None:
            found = { 'found': person, 's': 'bank' }

            return found

    # ipriziv
    if static.loginCheck:
        persons = get_persons()['data']
   
        for person in persons:
            
            if get_passport(person) == kod:
                found = { 'found': person, 's': 'ipriziv' }

    return found

import re

def capitalize_tajics(str):
    joins = ' '
    tajics = ['оглы','уулу','улы','угли']

    if str.find('-') != -1:
        joins = '-'

    str = re.split(' |-', str)

    b = []

    for a in str:
        if a.lower() in tajics:
            a = a.lower()
        else:
            a = a.capitalize()
        b.append(a)

    x = joins.join(a for a in b)

    return x

@app.post('/add-person')
def add_person():
    arg = json.loads(request.data)

    person = add_person_bd(
        arg['passport_serial'], 
        capitalize_tajics(arg['last_name']), 
        capitalize_tajics(arg['first_name']), 
        capitalize_tajics(arg['patronymic']), 
        arg['birth_date'], 
        arg['birth_place'], 
        arg['passport_issue'], 
        arg['passport_issue_date'], 
        arg['passport_division_code'], 
        arg['address'],
        arg['phone_home'],
        arg['recruitment_office_id'],
        arg['codeword']
    )

    count_add()

    return { 'reload': 'add', 'person': person }

@app.post('/edit-person')
def edit_person():
    arg = json.loads(request.data)

    person = edit_person_bd(
        arg['id'],
        arg['passport_serial'], 
        arg['last_name'],
        arg['first_name'], 
        arg['patronymic'], 
        arg['birth_date'], 
        arg['birth_place'], 
        arg['passport_issue'], 
        arg['passport_issue_date'], 
        arg['passport_division_code'], 
        arg['address'],
        arg['phone_home'],
        arg['recruitment_office_id'],
        arg['codeword']
    )

    count_add()

    return { 'reload': 'edit', 'person': person }

import openpyxl, shutil, os

@app.post('/export')
def export():
    arg = json.loads(request.data)

    persons_to_export = arg['data']

    shutil.copy(static.cfg["excel"]["formPath"], static.cfg["excel"]["copyFormPath"])
    
    workbook = openpyxl.load_workbook(static.cfg["excel"]["copyFormPath"])
    sheet = workbook['Лист1']

    row_start = 4
    index = 0

    for person in persons_to_export:
        birthDate = datetime.datetime.strptime(person['birthDate'],"%a, %d %b %Y %H:%M:%S GMT")
        passportIssueDate = datetime.datetime.strptime(person['passportIssueDate'],"%a, %d %b %Y %H:%M:%S GMT")

        sheet.cell(row=row_start, column=1).value = index # индексация с 0
        sheet.cell(row=row_start, column=2).value = person['lastName'] # фамилия
        sheet.cell(row=row_start, column=3).value = person['firstName'] # имя
        sheet.cell(row=row_start, column=4).value = person['patronymic'] # отчество
        sheet.cell(row=row_start, column=5).value = birthDate.strftime("%d.%m.%Y") # дата рождения
        sheet.cell(row=row_start, column=6).value = person['birthPlace'] # место рождения
        sheet.cell(row=row_start, column=7).value = (str)(person['passportSerial']).encode('cp1251') # серия и номер паспорта
        sheet.cell(row=row_start, column=8).value = person['passportIssue'] # где и когда выдан
        sheet.cell(row=row_start, column=9).value = passportIssueDate.strftime("%d.%m.%Y") # дата выдачи
        sheet.cell(row=row_start, column=10).value = person['passportDivisionCode'] # код подразделения
        sheet.cell(row=row_start, column=11).value = person['address'] # адрес проживания
        sheet.cell(row=row_start, column=12).value = 'Москва' # город
        sheet.cell(row=row_start, column=13).value = person['phoneMobile'] # домашний телефон
        sheet.cell(row=row_start, column=14).value = person['phoneMobile'] # моб телефон
        sheet.cell(row=row_start, column=15).value = person['codeword'] # код слово
        sheet.cell(row=row_start, column=16).value = 'Тимирязевский' # тимирязевский

        row_start = row_start + 1
        index = index + 1

    if check_local_database():
        name = ""
        k = 1
        now = datetime.datetime.now().strftime("%Y.%m.%d")

        ExcelServerPath = static.cfg["default"]["serverPath"]+"\\"+static.cfg["default"]["folder"]

        while (True):
            name = f"ВТБ_{now}_{k}.xlsx"
            if os.path.exists(ExcelServerPath+static.cfg["excel"]["savePath"]+name):
                k = k + 1
            else:
                break
        
        workbook.save(ExcelServerPath+static.cfg["excel"]["savePath"]+name)

        save_folder_path = ExcelServerPath+static.cfg["excel"]["exportsPath"]+"\\"+now+"\\"

        if not os.path.exists(save_folder_path):
            os.makedirs(save_folder_path)

        name = ""
        k = 1

        while (True):
            name = f"ВТБ_{now}_{k}.xlsx"
            if os.path.exists(save_folder_path+name):
                k = k + 1
            else:
                break 
        
        workbook.save(save_folder_path+name)
    else:
        workbook.save(static.cfg["excel"]["copyFormPath"])

    return { 'export': 'okay' }

total = static.get_total_count()

@app.get('/add/')
def add_get():
    if request.method == 'GET':
        return render_template('add.html', db_check=check_local_database(), totalCount=total, check=static.loginCheck )
    
@app.get('/add/<kod>')
def add_get_kod(kod):
    if request.method == 'GET':
        #print(kod)
        return render_template('add.html', db_check=check_local_database(), totalCount=total, check=static.loginCheck, edit_kod=kod)
    
@app.get('/base/')
def base():
    rqt = json.loads(request.args.get('request'))

    persons = json.dumps(get_persons_db(rqt['limit'],rqt['offset']))



    return "{ "+f"\"records\": {persons}, \"total\": {total_count()} "+" }"

if __name__ == '__main__':
    app.run(debug=True)